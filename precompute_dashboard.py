#!/usr/bin/env python3
"""Precompute all data for the sophisticated dashboard."""
import warnings, json
warnings.filterwarnings('ignore')
import numpy as np, pandas as pd
import openpyxl
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans, DBSCAN
from sklearn.metrics import silhouette_score, davies_bouldin_score, silhouette_samples
from sklearn.decomposition import PCA
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import cdist

np.random.seed(42)

PATH = "BaseDedatosFinal.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["Data Final"]
headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is not None:
        rows.append(list(row))
df = pd.DataFrame(rows, columns=headers)
dept = df['Departamento'].values

# Feature engineering (7 features as in the script)
df['Oficinas_por_100k'] = df['Num_Cajas'] / (df['PoblacionTotal_Censo2017'] / 100000)
df['Depositos_por_Capita'] = df['Depositos_Cajas_PEN_MM'] * 1e6 / df['Poblacion_18_70']
pea_total = df['PEA_ocupada_miles_2024'] + df['PEA_no_ocupada_miles_2024']
df['Tasa_Empleo'] = df['PEA_ocupada_miles_2024'] / pea_total
df['Internet_por_100k'] = df['Conexión_Internet_Fijo'] / (df['PoblacionTotal_Censo2017'] / 100000)
df['Movil_por_100k'] = df['NroLineasTelefoniaMovil'] / (df['PoblacionTotal_Censo2017'] / 100000)
df['PBI_por_Capita'] = df['PBI_miles_PEN'] * 1000 / df['PoblacionTotal_Censo2017']
df['PNP_por_100k'] = df['Num_PNP_2026'] / (df['PoblacionTotal_Censo2017'] / 100000)
df['Brecha_Inclusion'] = df['NBI_%_2024'] / (df['Oficinas_por_100k'] + 0.01)

FEATURES = ['Oficinas_por_100k','Depositos_por_Capita','NBI_%_2024','Ingreso_Prom_PEN_2024','Tasa_Empleo','PBI_por_Capita','Internet_por_100k']
FEAT_LABELS = ['Oficinas x100k','Depositos x adulto','NBI %','Ingreso Prom.','Tasa Empleo','PBI per capita','Internet x100k']

X = df[FEATURES].values
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)

# PCA
pca = PCA(n_components=3)
X_pca = pca.fit_transform(X_scaled)
pca_var = pca.explained_variance_ratio_
loadings = pca.components_

# K-Means for K=2..6
kmeans_results = {}
for k in range(2, 7):
    km = KMeans(n_clusters=k, random_state=42, n_init=50, max_iter=500)
    labels = km.fit_predict(X_scaled)
    sil = silhouette_score(X_scaled, labels)
    db = davies_bouldin_score(X_scaled, labels)
    sil_samples = silhouette_samples(X_scaled, labels).tolist()
    kmeans_results[k] = {
        'labels': [int(l) for l in labels],
        'centroids': km.cluster_centers_.tolist(),
        'silhouette': round(sil, 4),
        'davies_bouldin': round(db, 4),
        'silhouette_samples': [round(s, 4) for s in sil_samples],
        'inertia': round(km.inertia_, 2)
    }

# K=3 final model (used by the script)
KM3 = kmeans_results[3]
labels_3 = np.array(KM3['labels'])
dbscan = DBSCAN(eps=2.2, min_samples=2)
db_labels = dbscan.fit_predict(X_scaled)

# Per-cluster profiles for K=3
cluster_profiles = {}
for c in range(3):
    mask = labels_3 == c
    prof = {}
    for v in FEATURES:
        vals = df[v].values[mask]
        prof[v] = {'mean': round(float(vals.mean()), 4), 'std': round(float(vals.std()), 4)}
    prof['depts'] = dept[mask].tolist()
    cluster_profiles[c] = prof

# Silhouette per dept for K=3
sil_samples_3 = {str(dept[i]): KM3['silhouette_samples'][i] for i in range(24)}

# Data rows
data_rows = []
for i in range(24):
    row = {
        'd': str(dept[i]),
        'c': int(labels_3[i]),
        'db': int(db_labels[i]),
        'sil': KM3['silhouette_samples'][i],
        'nc': float(df['Num_Cajas'].iloc[i]),
        'dep': float(df['Depositos_Cajas_PEN_MM'].iloc[i]),
        'pob': float(df['PoblacionTotal_Censo2017'].iloc[i]),
        'ing': float(df['Ingreso_Prom_PEN_2024'].iloc[i]),
        'nbi': float(df['NBI_%_2024'].iloc[i]),
        'den': float(df['Denuncias_2024'].iloc[i]),
        'pnp': float(df['Num_PNP_2026'].iloc[i]),
        'pbi': float(df['PBI_miles_PEN'].iloc[i]),
        'of': float(df['Oficinas_por_100k'].iloc[i]),
        'dpc': float(df['Depositos_por_Capita'].iloc[i]),
        'te': float(df['Tasa_Empleo'].iloc[i]),
        'int': float(df['Internet_por_100k'].iloc[i]),
        'mov': float(df['Movil_por_100k'].iloc[i]),
        'pbi_pc': float(df['PBI_por_Capita'].iloc[i]),
        'pnp_100k': float(df['PNP_por_100k'].iloc[i]),
        'brecha': float(df['Brecha_Inclusion'].iloc[i]),
        'pc1': float(X_pca[i,0]), 'pc2': float(X_pca[i,1]), 'pc3': float(X_pca[i,2]),
    }
    data_rows.append(row)

# K selection metrics (for K=2..8)
k_range = list(range(2, 9))
metrics_k = {}
max_k_possible = min(9, X_scaled.shape[0])
for k in range(2, max_k_possible):
    km = KMeans(n_clusters=k, random_state=42, n_init=50, max_iter=500)
    lbls = km.fit_predict(X_scaled)
    metrics_k[k] = {
        'inertia': round(km.inertia_, 2),
        'silhouette': round(silhouette_score(X_scaled, lbls), 4),
        'davies_bouldin': round(davies_bouldin_score(X_scaled, lbls), 4)
    }

# Random Forest
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import cross_val_score
FEAT_RF = [f for f in FEATURES if f != 'Depositos_por_Capita']
X_rf = df[FEAT_RF].values
y_rf = df['Depositos_por_Capita'].values
rf = RandomForestRegressor(n_estimators=200, max_depth=5, random_state=42, n_jobs=-1)
rf.fit(X_rf, y_rf)
cv_scores = cross_val_score(rf, X_rf, y_rf, cv=5, scoring='r2')
rf_imp = {FEAT_RF[i]: round(float(rf.feature_importances_[i]), 4) for i in range(len(FEAT_RF))}

# Dendrogram linkage
Z = linkage(X_scaled, method='ward')

# Hopkins statistic
from sklearn.neighbors import NearestNeighbors
def hopkins(X):
    n = X.shape[0]
    d = X.shape[1]
    X_unif = np.random.uniform(X.min(axis=0), X.max(axis=0), size=(n, d))
    nbrs_real = NearestNeighbors(n_neighbors=2).fit(X)
    dist_real, _ = nbrs_real.kneighbors(X)
    u_real = dist_real[:, 1]
    nbrs_unif = NearestNeighbors(n_neighbors=2).fit(X_unif)
    dist_unif, _ = nbrs_unif.kneighbors(X_unif)
    u_unif = dist_unif[:, 1]
    H = u_unif.sum() / (u_real.sum() + u_unif.sum())
    return round(H, 4)
H = hopkins(X_scaled)

# DBSCAN info
db_outliers = dept[db_labels == -1].tolist()
db_clusters = len(set(db_labels)) - (1 if -1 in db_labels else 0)

# Build output JSON
OUT = {
    'depts': [str(d) for d in dept],
    'features': FEATURES,
    'feat_labels': FEAT_LABELS,
    'n_depts': 24,
    'n_features': 7,
    'pca_var': [round(v, 4) for v in pca_var],
    'pca_var_pct': [round(v*100, 1) for v in pca_var],
    'pca_total_2d': round(sum(pca_var[:2])*100, 1),
    'pca_total_3d': round(sum(pca_var)*100, 1),
    'loadings': [{'v': FEATURES[i], 'pc1': round(loadings[0,i],4), 'pc2': round(loadings[1,i],4), 'pc3': round(loadings[2,i],4)} for i in range(7)],
    'kmeans': {str(k): v for k, v in kmeans_results.items()},
    'cluster_profiles': {str(c): cluster_profiles[c] for c in range(3)},
    'dbscan': {'n_outliers': int((db_labels==-1).sum()), 'outliers': [str(d) for d in db_outliers], 'n_clusters': db_clusters},
    'rf_importance': rf_imp,
    'rf_cv_r2': round(float(cv_scores.mean()), 4),
    'rf_cv_std': round(float(cv_scores.std()), 4),
    'hopkins': H,
    'metrics_k': {str(k): v for k, v in metrics_k.items()},
    'data': data_rows,
    'linkage': [[round(float(z), 4) for z in row] for row in Z],
}

print(json.dumps(OUT, indent=2, ensure_ascii=False))
